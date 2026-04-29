"""
Lambda: genera_prospetto
GET /pratica/{id_pratica}/prospetto

Compila il template Excel del "Prospetto di applicazione del decreto di
ricostruzione carriera" sostituendo i placeholder [Campo] con i dati
estratti da DynamoDB (METADATA + DOCUMENTO#decreto_ricostruzione#...).

Risposta 200:
{
  "url": "https://s3.amazonaws.com/..."   ← presigned URL download (GET)
}

Struttura DynamoDB attesa per item decreto_ricostruzione:
  PK: "PRATICA#{id_pratica}"
  SK: "DOCUMENTO#decreto_ricostruzione#{uuid}"
  tipo_documento: "decreto_ricostruzione"

  dati_anagrafici:
    codice_fiscale, data_nascita, nome_cognome (o nome + cognome separati)

  dati_professionali:
    qualifica_funzionale
    data_decorrenza_giuridica
    data_decorrenza_economica
    classe_stipendiale
    data_scadenza_stipendi

  intestazione:
    numero_decreto, data_decreto, istituto_amministrante

  articolo_2:
    classe_stipendiale, data_scadenza
    periodo_totale_fini_giuridici_economici  (es. "X anni Y mesi Z giorni")
    periodo_totale_soli_fini_economici

  articolo_4 (o lista assegni):
    codice_assegno, importo_assegno_ad_personam, numero_mensilita
    data_decorrenza_assegno, data_scadenza_assegno

  visti (lista o dict):
    numero_visto, data_visto

  assenze (lista):
    data_inizio (o periodo_fruizione_inizio), data_fine (o periodo_fruizione_fine)
    tipologia_assenza

  METADATA:
    prescrizione: "Sì" / "No"
    data_prescrizione

Env vars:
  DYNAMODB_TABLE       – default: RTS_Pratiche
  REGION               – default: eu-central-1
  TEMPLATE_BUCKET      – bucket S3 con template Excel  (OBBLIGATORIO)
  TEMPLATE_KEY         – key S3 del template           (default: templates/prospetto_ricostruzione.xlsx)
  OUTPUT_BUCKET        – bucket S3 output              (default: TEMPLATE_BUCKET)
  OUTPUT_PREFIX        – prefix S3 output              (default: prospetti)
  PRESIGNED_URL_EXPIRY – secondi validità URL          (default: 3600)
"""

import calendar
import io
import json
import logging
import os
import re
from datetime import date, datetime, timedelta

import boto3
from boto3.dynamodb.conditions import Key
from botocore.exceptions import ClientError
from openpyxl import load_workbook

logger = logging.getLogger()
logger.setLevel(logging.INFO)

REGION = os.environ.get("REGION", "eu-central-1")
DYNAMODB_TABLE = os.environ.get("DYNAMODB_TABLE", "RTS_Pratiche")
TEMPLATE_BUCKET = os.environ["TEMPLATE_BUCKET"]
TEMPLATE_KEY = os.environ.get("TEMPLATE_KEY", "templates/prospetto_ricostruzione.xlsx")
OUTPUT_BUCKET = os.environ.get("OUTPUT_BUCKET", TEMPLATE_BUCKET)
OUTPUT_PREFIX = os.environ.get("OUTPUT_PREFIX", "prospetti")
PRESIGNED_URL_EXPIRY = int(os.environ.get("PRESIGNED_URL_EXPIRY", "3600"))

dynamodb = boto3.resource("dynamodb", region_name=REGION)
s3 = boto3.client("s3", region_name=REGION)

CORS_HEADERS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Headers": "Content-Type",
    "Access-Control-Allow-Methods": "GET, OPTIONS",
    "Content-Type": "application/json",
}

# ---------------------------------------------------------------------------
# Calcolo data di scadenza stipendi
# ---------------------------------------------------------------------------

# Anni totali di servizio (preruolo + in-ruolo) necessari per entrare in ogni classe.
# La chiave è la classe CORRENTE; il valore è la soglia anni per la classe SUCCESSIVA.
SCATTI_STIPENDIALI: dict[str, int | None] = {
    "00": 9,
    "09": 15,
    "15": 21,
    "21": 28,
    "28": 35,
    "35": None,  # ultima classe — nessuna scadenza
}


def _parse_date_it(s: str | None) -> date | None:
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def _add_period(d: date, anni: int, mesi: int, giorni: int) -> date:
    """Somma (o sottrae se negativi) anni/mesi/giorni a una data."""
    # Prima gestisci anni e mesi
    y = d.year + anni
    m = d.month + mesi
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    
    # Clamp day al massimo del mese (gestisce mesi con diversi giorni)
    max_day = calendar.monthrange(y, m)[1]
    day = min(d.day, max_day)
    
    # Crea la data con anno/mese corretti
    result = date(y, m, day)
    
    # Poi aggiungi/sottrai i giorni usando timedelta (gestisce automaticamente i cambi di mese)
    if giorni != 0:
        result = result + timedelta(days=giorni)
    
    return result


def _next_1_settembre(d: date) -> date:
    """Restituisce il 01/09 in corso o del prossimo anno rispetto a d."""
    candidate = date(d.year, 9, 1)
    return candidate if candidate >= d else date(d.year + 1, 9, 1)


_CLASSI_ORDINE = ["00", "09", "15", "21", "28", "35"]
# Soglia anni totali di servizio per USCIRE da ogni classe (entrare nella successiva)
_SOGLIE_ANNI = [9, 15, 21, 28, 35, None]


def _effective_start(data_decorrenza_economica: str | None, preruolo: dict | None) -> date | None:
    assunzione = _parse_date_it(data_decorrenza_economica)
    if not assunzione:
        return None
    p = preruolo or {}
    return _add_period(assunzione, -int(p.get("anni") or 0), -int(p.get("mesi") or 0), -int(p.get("giorni") or 0))


def calcola_righe_variazione_stipendi(
    data_decorrenza_economica: str | None,
    data_decorrenza_giuridica: str | None,
    preruolo_soli_fini_economici: dict | None,
    classe_stipendiale: str | None,
    qualifica_noiipa: str,
) -> list[dict]:
    """
    Genera una riga per ogni classe da "00" fino alla classe corrente.

    Regole scadenza per ogni riga i:
      target_i  = effective_start + _SOGLIE_ANNI[i]   (anni totali per uscire dalla classe i)
      scadenza_i = _next_1_settembre(target_i)
      Se scadenza_i <= data_decorrenza_economica → scadenza_i = data_decorrenza_economica
        (la transizione è già avvenuta al momento dell'assunzione in ruolo)

    Prox variazione automatica:
      "No" per tutte le righe precedenti l'ultima, "Sì" per l'ultima.

    Restituisce lista di dict con chiavi:
      dec_econ, dec_giur, scadenza, qualifica, classe, prox_var
    """
    classe_norm = _normalizza_classe_stipendiale(classe_stipendiale)
    if classe_norm not in _CLASSI_ORDINE:
        classe_norm = "00"

    idx_corrente = _CLASSI_ORDINE.index(classe_norm)
    assunzione = _parse_date_it(data_decorrenza_economica)
    eff_start = _effective_start(data_decorrenza_economica, preruolo_soli_fini_economici)
    dec_econ = data_decorrenza_economica or ""
    dec_giur = data_decorrenza_giuridica or ""

    rows = []
    for i in range(idx_corrente + 1):
        soglia = _SOGLIE_ANNI[i]
        if eff_start and soglia:
            target = _add_period(eff_start, soglia, 0, 0)
            scadenza_d = _next_1_settembre(target)
            # Se la transizione è già avvenuta prima/uguale all'assunzione → usa data assunzione
            if assunzione and scadenza_d <= assunzione:
                scadenza_str = assunzione.strftime("%d/%m/%Y")
            else:
                scadenza_str = scadenza_d.strftime("%d/%m/%Y")
        else:
            scadenza_str = ""

        rows.append({
            "dec_econ":   dec_econ,
            "dec_giur":   dec_giur,
            "scadenza":   scadenza_str,
            "qualifica":  qualifica_noiipa,
            "classe":     _CLASSI_ORDINE[i],
            "prox_var":   "Sì" if i == idx_corrente else "No",
        })

    return rows


def calcola_data_scadenza_stipendi(
    data_decorrenza_economica: str | None,
    preruolo_soli_fini_economici: dict | None,
    classe_stipendiale: str | None,
) -> str:
    """Scadenza della classe corrente (ultima riga della variazione stipendi)."""
    classe_norm = _normalizza_classe_stipendiale(classe_stipendiale)
    soglia_anni = SCATTI_STIPENDIALI.get(classe_norm)
    if soglia_anni is None:
        return ""

    eff_start = _effective_start(data_decorrenza_economica, preruolo_soli_fini_economici)
    if not eff_start:
        return ""

    target = _add_period(eff_start, soglia_anni, 0, 0)
    return _next_1_settembre(target).strftime("%d/%m/%Y")


# ---------------------------------------------------------------------------
# Mapping classe stipendiale: testo → codice numerico
# ---------------------------------------------------------------------------
#
# Converte descrizioni testuali della posizione stipendiale (estratte dall'LLM)
# nei codici numerici corrispondenti (00, 09, 15, 21, 28, 35).
#
_CLASSE_STIPENDIALE_MAPPING: dict[str, str] = {
    # Posizioni stipendiali
    "prima posizione stipendiale": "00",
    "prima posizione": "00",
    "1a posizione stipendiale": "00",
    "1a posizione": "00",
    "1° posizione stipendiale": "00",
    "1° posizione": "00",
    "1^ posizione stipendiale": "00",
    "1^ posizione": "00",
    "i posizione stipendiale": "00",
    "i posizione": "00",
    
    "seconda posizione stipendiale": "09",
    "seconda posizione": "09",
    "2a posizione stipendiale": "09",
    "2a posizione": "09",
    "2° posizione stipendiale": "09",
    "2° posizione": "09",
    "2^ posizione stipendiale": "09",
    "2^ posizione": "09",
    "ii posizione stipendiale": "09",
    "ii posizione": "09",
    
    "terza posizione stipendiale": "15",
    "terza posizione": "15",
    "3a posizione stipendiale": "15",
    "3a posizione": "15",
    "3° posizione stipendiale": "15",
    "3° posizione": "15",
    "3^ posizione stipendiale": "15",
    "3^ posizione": "15",
    "iii posizione stipendiale": "15",
    "iii posizione": "15",
    
    "quarta posizione stipendiale": "21",
    "quarta posizione": "21",
    "4a posizione stipendiale": "21",
    "4a posizione": "21",
    "4° posizione stipendiale": "21",
    "4° posizione": "21",
    "4^ posizione stipendiale": "21",
    "4^ posizione": "21",
    "iv posizione stipendiale": "21",
    "iv posizione": "21",
    
    "quinta posizione stipendiale": "28",
    "quinta posizione": "28",
    "5a posizione stipendiale": "28",
    "5a posizione": "28",
    "5° posizione stipendiale": "28",
    "5° posizione": "28",
    "5^ posizione stipendiale": "28",
    "5^ posizione": "28",
    "v posizione stipendiale": "28",
    "v posizione": "28",
    
    "sesta posizione stipendiale": "35",
    "sesta posizione": "35",
    "6a posizione stipendiale": "35",
    "6a posizione": "35",
    "6° posizione stipendiale": "35",
    "6° posizione": "35",
    "6^ posizione stipendiale": "35",
    "6^ posizione": "35",
    "vi posizione stipendiale": "35",
    "vi posizione": "35",
}


def _normalizza_classe_stipendiale(classe: str | None) -> str:
    """
    Converte la descrizione testuale della classe stipendiale nel codice numerico.
    Se il valore è già un codice numerico (00-35), lo restituisce normalizzato (con zero-padding).
    Se il valore è testo (es. "prima posizione stipendiale"), lo converte al codice corrispondente.
    Se nessun match, restituisce "00" come default.
    
    Args:
        classe: Testo o codice della classe stipendiale
        
    Returns:
        Codice numerico zero-padded (es. "00", "09", "15", "21", "28", "35")
    """
    if not classe:
        return "00"
    
    classe_str = str(classe).strip()
    
    # Se è già un numero, normalizza con zero-padding
    if classe_str.isdigit():
        return classe_str.zfill(2)
    
    # Altrimenti cerca nel mapping testuale
    classe_lower = classe_str.lower()
    
    # Cerca match esatto
    if classe_lower in _CLASSE_STIPENDIALE_MAPPING:
        return _CLASSE_STIPENDIALE_MAPPING[classe_lower]
    
    # Cerca match parziale (contiene una delle chiavi)
    for key, code in _CLASSE_STIPENDIALE_MAPPING.items():
        if key in classe_lower:
            return code
    
    # Default: prima posizione
    logger.warning("Classe stipendiale non riconosciuta: '%s', uso default '00'", classe_str)
    return "00"


# ---------------------------------------------------------------------------
# Mapping qualifica funzionale → codice NoiPA
# ---------------------------------------------------------------------------
#
# Ogni entry: (keywords_da_cercare_nel_testo, codice, descrizione_ufficiale_noiipa)
# Le keyword vengono cercate nell'ordine → metti le più specifiche prima.
# Il match avviene se TUTTE le keyword di un gruppo sono presenti nel testo normalizzato.
#
_NOIIPA_RULES: list[tuple[list[str], str, str]] = [
    # ---- AFAM (prima dei generici "assistente" / "funzionario") ----
    (["afam", "elevata qualificazione", "ex kc19"],  "KC26", "AFAM Elevata qualificazione ex KC19"),
    (["afam", "elevata qualificazione"],             "KC16", "AFAM Elevata Qualificazione"),
    (["afam", "funzionario"],                        "KC17", "AFAM Funzionario"),
    (["afam", "assistente"],                         "KC43", "AFAM - Assistente"),
    (["afam", "operatore"],                          "KC41", "AFAM - Operatore"),
    # ---- Docenti (più specifici prima) ----
    (["diplomato", "secondari", "ii grado"],         "KA06", "Docente diplomato istituti secondari II grado"),
    (["diplomato", "superior"],                      "KA06", "Docente diplomato istituti secondari II grado"),
    (["laureato", "secondari", "ii grado"],          "KA08", "Docente laureato istituti secondari II grado"),
    (["laureato", "superior"],                       "KA08", "Docente laureato istituti secondari II grado"),
    (["secondari", "ii grado"],                      "KA06", "Docente diplomato istituti secondari II grado"),
    (["superior"],                                   "KA06", "Docente diplomato istituti secondari II grado"),
    (["materna"],                                    "KA05", "Docente scuola materna ed elementare"),
    (["infanzia"],                                   "KA05", "Docente scuola materna ed elementare"),
    (["elementar"],                                  "KA05", "Docente scuola materna ed elementare"),
    (["primaria"],                                   "KA05", "Docente scuola materna ed elementare"),
    (["media"],                                      "KA07", "Docente scuola media"),
    (["secondaria", "i grado"],                      "KA07", "Docente scuola media"),
    (["secondari"],                                  "KA06", "Docente diplomato istituti secondari II grado"),
    # ---- Capi istituto ----
    (["capo", "istituto"],                           "KA11", "Capi di istituto che non hanno acquisito la qualifica di dirigenti scolastici"),
    # ---- ATA ----
    (["collaborator"],                               "KA41", "Collaboratori"),
    (["operatore"],                                  "KA42", "Operatori"),
    (["assistente"],                                 "KA43", "Assistenti"),
    (["dsga"],                                       "KA12", "Funzionari ed elevata qualificazione"),
    (["direttore", "servizi", "generali"],           "KA12", "Funzionari ed elevata qualificazione"),
    # ---- Funzionari (KA12 prima di KA13: "funzionari" plurale → KA12) ----
    (["funzionari", "elevata"],                      "KA12", "Funzionari ed elevata qualificazione"),
    (["funzionario"],                                "KA12", "Funzionari ed elevata qualificazione"),
    (["dirigente"],                                  "KA12", "Funzionari ed elevata qualificazione"),
    (["elevata qualificazione"],                     "KA13", "Funz. elevata qualificazione"),
]


def _noiipa_qualifica(qualifica: str | None) -> str:
    """
    Mappa il testo libero della qualifica funzionale al codice NoiPA ufficiale.
    Restituisce "CODICE - Descrizione ufficiale NoiPA".
    Se nessun match, restituisce il testo originale.
    """
    if not qualifica:
        return ""
    lower = qualifica.lower().strip()
    for keywords, code, _ in _NOIIPA_RULES:
        if all(kw in lower for kw in keywords):
            return code
    return qualifica  # nessun match: testo grezzo dal decreto


def calcola_righe_variazione_assegni(
    assegni_raw: list | dict,
    data_decorrenza_economica: str | None,
    data_scadenza_stipendi: str,
) -> list[dict]:
    """
    Genera le righe per la variazione assegni secondo le regole:
    - Per ogni assegno, crea DUE righe:
      1. Cessazione: stesso codice assegno, data scadenza = data_scadenza_stipendi,
         tipo operazione "Cessazione", Prox variazione automatica "No"
      2. Inserimento: stesso codice assegno, data scadenza = data_scadenza_stipendi,
         tipo operazione "Inserimento", Prox variazione automatica "Sì"
    
    La data_decorrenza viene impostata a data_decorrenza_economica.
    La data_scadenza viene impostata a data_scadenza_stipendi per entrambe le righe.
    
    Restituisce lista di dict con chiavi:
      data_decorrenza, codice_assegno, importo, mensilita, data_scadenza, tipo_op, prox_var
    """
    # Normalizza input assegni_raw in lista
    if isinstance(assegni_raw, dict):
        assegni_list = [assegni_raw]
    elif isinstance(assegni_raw, list):
        assegni_list = assegni_raw
    else:
        assegni_list = []
    
    if not assegni_list:
        return []
    
    rows = []
    for assegno in assegni_list:
        if not isinstance(assegno, dict):
            continue
        
        codice = _get(assegno, "codice_assegno") or _get(assegno, "codice") or ""
        importo = _get(assegno, "importo_assegno_ad_personam") or _get(assegno, "importo") or ""
        mensilita = _get(assegno, "numero_mensilita") or _get(assegno, "natura_importo") or ""
        
        # La data_decorrenza è sempre quella economica
        data_dec = data_decorrenza_economica or ""
        
        # La data_scadenza è sempre quella della variazione stipendi e assegni
        data_scad = data_scadenza_stipendi or ""
        
        # Prima riga: Cessazione
        rows.append({
            "data_decorrenza": data_dec,
            "codice_assegno": codice,
            "importo": importo,
            "mensilita": mensilita,
            "data_scadenza": data_scad,
            "tipo_op": "Cessazione",
            "prox_var": "No",
        })
        
        # Seconda riga: Inserimento
        rows.append({
            "data_decorrenza": data_dec,
            "codice_assegno": codice,
            "importo": importo,
            "mensilita": mensilita,
            "data_scadenza": data_scad,
            "tipo_op": "Inserimento",
            "prox_var": "Sì",
        })
    
    return rows


# ---------------------------------------------------------------------------
# Helpers DynamoDB
# ---------------------------------------------------------------------------

def _get(d: dict | None, *keys, default=""):
    for k in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(k)
        if d is None:
            return default
    return d if d is not None else default


def _first(*values, default=""):
    for v in values:
        if v:
            return v
    return default


def query_metadata(id_pratica: str) -> dict:
    table = dynamodb.Table(DYNAMODB_TABLE)
    resp = table.get_item(Key={"PK": f"PRATICA#{id_pratica}", "SK": "METADATA"})
    return resp.get("Item") or {}


def query_documento_decreto(id_pratica: str) -> dict | None:
    """Restituisce il primo item decreto_ricostruzione trovato."""
    table = dynamodb.Table(DYNAMODB_TABLE)
    kwargs = {
        "KeyConditionExpression": (
            Key("PK").eq(f"PRATICA#{id_pratica}")
            & Key("SK").begins_with("DOCUMENTO#decreto_ricostruzione#")
        )
    }
    items = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        last = resp.get("LastEvaluatedKey")
        if not last:
            break
        kwargs["ExclusiveStartKey"] = last
    # Prefer the most recent extraction
    if not items:
        return None
    items.sort(key=lambda x: x.get("extractedAt", ""), reverse=True)
    return items[0]


def query_documento_visto(id_pratica: str) -> dict | None:
    """Restituisce il primo item visto trovato."""
    table = dynamodb.Table(DYNAMODB_TABLE)
    kwargs = {
        "KeyConditionExpression": (
            Key("PK").eq(f"PRATICA#{id_pratica}")
            & Key("SK").begins_with("DOCUMENTO#visto#")
        )
    }
    items = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        last = resp.get("LastEvaluatedKey")
        if not last:
            break
        kwargs["ExclusiveStartKey"] = last
    # Prefer the most recent extraction
    if not items:
        return None
    items.sort(key=lambda x: x.get("extractedAt", ""), reverse=True)
    return items[0]


# ---------------------------------------------------------------------------
# Data extraction helpers
# ---------------------------------------------------------------------------

def _extract_nome_cognome(decreto: dict) -> tuple[str, str]:
    """
    Restituisce (cognome, nome) dall'item decreto.
    Cerca prima campi separati, poi divide nome_cognome (formato COGNOME NOME).
    """
    anag = decreto.get("dati_anagrafici") or {}
    cognome = _get(anag, "cognome") or _get(decreto, "cognome")
    nome = _get(anag, "nome") or _get(decreto, "nome")
    if cognome and nome:
        return str(cognome), str(nome)

    nome_cognome = (
        _get(anag, "nome_cognome")
        or _get(decreto, "nome_cognome")
        or ""
    )
    if nome_cognome:
        parts = nome_cognome.split()
        if len(parts) >= 2:
            # Convenzione italiana: COGNOME NOME
            return parts[0], " ".join(parts[1:])
        return nome_cognome, ""
    return cognome or "", nome or ""


def _format_assenza(a: dict) -> str:
    inizio = (
        a.get("periodo_fruizione_inizio")
        or a.get("data_inizio_assenza")
        or a.get("data_inizio")
        or ""
    )
    fine = (
        a.get("periodo_fruizione_fine")
        or a.get("data_fine_assenza")
        or a.get("data_fine")
        or ""
    )
    tipo = a.get("tipologia_assenza") or a.get("tipo_assenza") or a.get("tipo") or ""
    if inizio and fine:
        base = f"dal {inizio} al {fine}"
    elif inizio:
        base = f"dal {inizio}"
    else:
        return tipo or ""
    return f"{base} ({tipo})" if tipo else base


def _format_anzianita(val) -> str:
    """Formatta un valore di anzianità come 'X anni Y mesi Z giorni' o stringa raw."""
    if isinstance(val, dict):
        anni = val.get("anni", 0) or 0
        mesi = val.get("mesi", 0) or 0
        giorni = val.get("giorni", 0) or 0
        parts = []
        if anni:
            parts.append(f"{anni} ann{'o' if anni == 1 else 'i'}")
        if mesi:
            parts.append(f"{mesi} mes{'e' if mesi == 1 else 'i'}")
        if giorni:
            parts.append(f"{giorni} giorn{'o' if giorni == 1 else 'i'}")
        return ", ".join(parts) if parts else "0 giorni"
    return str(val) if val else ""



def _soggetto_a_prescrizione(metadata: dict) -> str:
    prescrizione = metadata.get("prescrizione", "No")
    if str(prescrizione).strip().lower() in ("sì", "si", "s", "yes", "true", "1"):
        data = metadata.get("data_prescrizione") or ""
        return f"con prescrizione al {data}".strip() if data else "con prescrizione"
    return ""


# ---------------------------------------------------------------------------
# Placeholder replacement
# ---------------------------------------------------------------------------

def _replace_all(text: str, mapping: dict) -> str:
    """Sostituisce tutti i placeholder [key] nel testo con i valori del mapping."""
    for placeholder, value in mapping.items():
        text = text.replace(f"[{placeholder}]", str(value) if value is not None else "")
    return text


def _build_placeholder_map(decreto: dict, metadata: dict, visto: dict | None = None) -> dict:
    anag = decreto.get("dati_anagrafici") or {}
    prof = decreto.get("dati_professionali") or {}
    int_ = decreto.get("intestazione") or {}
    art2 = decreto.get("articolo_2") or {}
    art4 = decreto.get("articolo_4") or {}

    # Supporta sia singolo dict che lista per art4
    if isinstance(art4, list):
        art4 = art4[0] if art4 else {}

    assenze_raw = decreto.get("assenze") or []

    cognome, nome = _extract_nome_cognome(decreto)

    codice_fiscale = (
        _get(anag, "codice_fiscale")
        or _get(decreto, "codice_fiscale")
        or metadata.get("codice_fiscale", "")
    )
    data_nascita = _get(anag, "data_nascita") or _get(decreto, "data_nascita") or ""

    qualifica_raw = (
        _get(prof, "qualifica_funzionale")
        or _get(prof, "qualifica")
        or _get(decreto, "qualifica_funzionale")
        or ""
    )

    data_nomina_in_ruolo = _get(prof, "data_decorrenza_giuridica") or ""
    data_decorrenza_giuridica = _get(prof, "data_decorrenza_giuridica") or ""
    data_decorrenza_economica = _first(
        _get(prof, "data_decorrenza_economica"),
        _get(art2, "data_decorrenza_economica"),
    )
    classe_stipendiale_raw = (
        _get(art2, "classe_stipendiale")
        or _get(prof, "classe_stipendiale")
        or _get(decreto, "classe_stipendiale")
        or ""
    )
    # Normalizza la classe stipendiale (da testo a codice numerico)
    classe_stipendiale = _normalizza_classe_stipendiale(classe_stipendiale_raw)

    # Preruolo riconosciuto ai fini giuridici ed economici (per calcolo scadenza stipendi)
    periodo_giuridico_economico = (
        _get(art2, "periodo_totale_fini_giuridici_economici")
        or _get(art2, "anzianita_fini_giuridici_economici")
        or _get(decreto, "periodo_totale_fini_giuridici_economici")
        or {}
    )
    if not isinstance(periodo_giuridico_economico, dict):
        periodo_giuridico_economico = {}
    
    # Preruolo riconosciuto ai soli fini economici (per documentazione)
    preruolo_soli_fini_economici = (
        _get(art2, "periodo_totale_soli_fini_economici")
        or _get(decreto, "preruolo_soli_fini_economici")
        or _get(decreto, "periodo_totale_soli_fini_economici")
        or {}
    )
    if not isinstance(preruolo_soli_fini_economici, dict):
        preruolo_soli_fini_economici = {}

    # Data scadenza stipendi: calcolata sulla base di decorrenza economica + anzianità giuridica-economica
    # Fallback al valore estratto dal decreto se disponibile
    data_scadenza_stipendi = calcola_data_scadenza_stipendi(
        data_decorrenza_economica=data_decorrenza_economica,
        preruolo_soli_fini_economici=periodo_giuridico_economico,
        classe_stipendiale=classe_stipendiale,
    ) or _get(art2, "data_scadenza") or _get(prof, "data_scadenza_stipendi") or ""
    data_conferma_in_ruolo = (
        _get(art2, "data_conferma_in_ruolo")
        or _get(prof, "data_conferma_in_ruolo")
        or _get(prof, "data_conferma_ruolo")
        or ""
    )

    numero_decreto = _get(int_, "numero_decreto") or _get(decreto, "numero_decreto") or ""
    data_decreto = _get(int_, "data_decreto") or _get(decreto, "data_decreto") or ""
    istituto = (
        _get(int_, "istituto_amministrante")
        or _get(int_, "istituto")
        or _get(decreto, "istituto_amministrante")
        or ""
    )

    codice_assegno = _get(art4, "codice_assegno") or _get(art4, "codice") or ""
    importo_assegno = (
        _get(art4, "importo_assegno_ad_personam")
        or _get(art4, "importo")
        or ""
    )
    numero_mensilita = (
        _get(art4, "numero_mensilita")
        or _get(art4, "natura_importo")
        or ""
    )
    data_decorrenza_assegno = (
        _get(art4, "data_decorrenza")
        or _get(art4, "data_decorrenza_assegno")
        or data_decorrenza_economica
        or ""
    )
    # Scadenza assegno: esplicita nel decreto (es. "01/01/9999" per non-riassorbibile)
    # altrimenti coincide con la scadenza stipendi
    data_scadenza_assegno = (
        _get(art4, "data_scadenza")
        or _get(art4, "data_scadenza_assegno")
        or data_scadenza_stipendi
        or ""
    )

    # Visti: recuperati dal documento visto separato
    numero_visto = _get(visto, "numero_visto") or _get(visto, "numero") or ""
    data_visto = _get(visto, "data_visto") or _get(visto, "data") or ""

    periodo_giur_econ = _format_anzianita(
        _get(art2, "periodo_totale_fini_giuridici_economici")
        or _get(art2, "anzianita_fini_giuridici_economici")
        or _get(decreto, "periodo_totale_fini_giuridici_economici")
    )
    periodo_solo_econ = _format_anzianita(
        _get(art2, "periodo_totale_soli_fini_economici")
        or _get(art2, "anzianita_soli_fini_economici")
        or _get(decreto, "periodo_totale_soli_fini_economici")
    )

    soggetto_prescriz = _soggetto_a_prescrizione(metadata)
    qualifica_noiipa = _noiipa_qualifica(qualifica_raw)

    # Righe variazione stipendi: una per ogni classe da "00" alla classe corrente
    righe_stipendi = calcola_righe_variazione_stipendi(
        data_decorrenza_economica=data_decorrenza_economica,
        data_decorrenza_giuridica=data_decorrenza_giuridica,
        preruolo_soli_fini_economici=periodo_giuridico_economico,
        classe_stipendiale=classe_stipendiale,
        qualifica_noiipa=qualifica_noiipa,
    )
    # data_scadenza_stipendi = scadenza dell'ultima riga (classe corrente)
    data_scadenza_stipendi = righe_stipendi[-1]["scadenza"] if righe_stipendi else (
        calcola_data_scadenza_stipendi(data_decorrenza_economica, periodo_giuridico_economico, classe_stipendiale)
        or _get(art2, "data_scadenza") or _get(prof, "data_scadenza_stipendi") or ""
    )
    
    # Righe variazione assegni: doppia riga (Cessazione + Inserimento) per ogni assegno
    # NON recuperare date dai metadati del decreto (articolo_4)
    # Usa sempre data_decorrenza_economica e data_scadenza_stipendi
    assegni_raw_list = decreto.get("articolo_4") or []
    righe_assegni = calcola_righe_variazione_assegni(
        assegni_raw=assegni_raw_list,
        data_decorrenza_economica=data_decorrenza_economica,
        data_scadenza_stipendi=data_scadenza_stipendi,
    )

    return {
        # Anagrafica
        "Codice Fiscale": codice_fiscale,
        "Data di nascita": data_nascita,
        "Cognome": cognome,
        "Nome": nome,
        # Date professionali
        "Data di nomina in ruolo": data_nomina_in_ruolo,
        "Data di decorrenza giuridica": data_decorrenza_giuridica,
        "Data di decorrenza giuridca": data_decorrenza_giuridica,   # typo nel template originale
        "Data di decorrenza economica": data_decorrenza_economica,
        # [Data di scadenza] in B27 = scadenza ultima classe; in B34 = assegni → gestiti separatamente
        "Data di scadenza": data_scadenza_stipendi,
        # Stipendi (usati solo se B27 ha una sola riga / fallback)
        "Qualifica professionale": qualifica_noiipa,
        "Classe stipendiale": classe_stipendiale,
        # Assegni
        "Data decorrenza": data_decorrenza_assegno,
        "Codice assegno": codice_assegno,
        "Importo assegno ad personam": importo_assegno,
        "Numero mensilità": numero_mensilita,
        # Annotazioni
        "numero del decreto": numero_decreto,
        "data del decreto": data_decreto,
        "Istituto dell'amministrato": istituto,
        "soggetto a prescrizione": soggetto_prescriz,
        "numero del visto": numero_visto,
        "data del visto": data_visto,
        "Data conferma in ruolo": data_conferma_in_ruolo,
        "Periodo totale anzianità ai fini giuridici ed economici": periodo_giur_econ,
        "Periodo totale anzianità ai soli fini economici": periodo_solo_econ,
        # alias tabella art. 2
        "Periodo totale di anzianità a fini giuridici ed economici": periodo_giur_econ,
        "Periodo totale di anzianità ai soli fini economici": periodo_solo_econ,
        # campi privati: gestiti direttamente in fill_excel
        "_assenze": assenze_raw,
        "_data_scadenza_assegno": data_scadenza_assegno,
        "_righe_stipendi": righe_stipendi,
        "_righe_assegni": righe_assegni,
    }


def _fill_assenze_in_cell(text: str, assenze: list[dict]) -> str:
    """
    Sostituisce i placeholder '1. [Durata assenza]', '2. [Durata assenza]', '3. [Durata assenza]'
    con i dati reali delle assenze.
    Se ci sono più assenze dei placeholder disponibili, le mostra tutte nell'ultimo placeholder.
    """
    num_placeholders = 0
    # Conta quanti placeholder ci sono nel testo
    for i in range(1, 100):
        if f"{i}. [Durata assenza]" in text:
            num_placeholders = i
        else:
            break
    
    if num_placeholders == 0:
        return text
    
    # Sostituisci i placeholder con le assenze
    for i in range(1, num_placeholders + 1):
        idx = i - 1
        placeholder = f"{i}. [Durata assenza]"
        
        if i < num_placeholders:
            # Placeholder intermedi: una assenza per placeholder
            if idx < len(assenze):
                value = f"{i}. {_format_assenza(assenze[idx])}"
            else:
                value = f"{i}."
        else:
            # Ultimo placeholder: metti tutte le assenze rimanenti
            if idx < len(assenze):
                remaining = []
                for j in range(idx, len(assenze)):
                    remaining.append(f"{j+1}. {_format_assenza(assenze[j])}")
                value = "\n".join(remaining)
            else:
                value = f"{i}."
        
        text = text.replace(placeholder, value)
    
    return text


# ---------------------------------------------------------------------------
# Excel template filling
# ---------------------------------------------------------------------------

SHEET_NAME = "Prospetto applicazione"


def _build_b27(header_line: str, righe: list[dict]) -> str:
    """
    Costruisce il contenuto testuale di B27 con header + una riga per classe.

    Formato dati (pipe-separated, allineato allo stile del template):
      dec_econ  | dec_giur  | scadenza  | qualifica  | classe  | prox_var
    """
    data_lines = []
    for r in righe:
        line = (
            f"{r['dec_econ']:<38}"
            f"| {r['dec_giur']:<40}"
            f"| {r['scadenza']:<34}"
            f"| {r['qualifica']:<50}"
            f"| {r['classe']:<44}"
            f"| {r['prox_var']}"
        )
        data_lines.append(line)
    return header_line + "\n" + "\n".join(data_lines)


def _build_b34(header_line: str, righe: list[dict]) -> str:
    """
    Costruisce il contenuto testuale di B34 con header + righe di variazione assegni.

    Formato dati (pipe-separated) allineato al template:
      Data Decorrenza | Scadenza | Tipo operazione | Codice | Classe | Prox variazione automatica | Importo A.L. | Natura importo
    """
    data_lines = []
    for r in righe:
        # Allinea le colonne al formato del template Excel
        line = (
            f"{r['data_decorrenza']:<38}"
            f"| {r['data_scadenza']:<40}"
            f"| {r['tipo_op']:<34}"
            f"| {r['codice_assegno']:<50}"
            f"| {r.get('classe', ''):<44}"
            f"| {r['prox_var']:<50}"
            f"|{r['importo']:<60}"
            f"|{r['mensilita']}"
        )
        data_lines.append(line)
    return header_line + "\n" + "\n".join(data_lines)


def fill_excel(template_bytes: bytes, mapping: dict) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' non trovato nel template")

    ws = wb[SHEET_NAME]
    assenze = mapping.pop("_assenze", [])
    data_scadenza_assegno = mapping.pop("_data_scadenza_assegno", "")
    righe_stipendi = mapping.pop("_righe_stipendi", [])
    righe_assegni = mapping.pop("_righe_assegni", [])

    # Mapping specifico per B34: [Data di scadenza] = scadenza assegno
    mapping_assegni = dict(mapping)
    if data_scadenza_assegno:
        mapping_assegni["Data di scadenza"] = data_scadenza_assegno

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            coord = cell.coordinate
            new_value = cell.value

            # Assenze (B4)
            new_value = _fill_assenze_in_cell(new_value, assenze)

            # B34: gestisce righe multiple assegni
            if coord == "B34":
                if righe_assegni:
                    lines = new_value.split("\n")
                    header = lines[0]  # intestazione colonne
                    new_value = _build_b34(header, righe_assegni)
                else:
                    # Fallback: usa il mapping con scadenza assegno
                    new_value = _replace_all(new_value, mapping_assegni)
            else:
                new_value = _replace_all(new_value, mapping)

            # B27: sostituisce la riga dati template con una riga per ogni classe
            if coord == "B27" and righe_stipendi:
                lines = new_value.split("\n")
                header = lines[0]  # intestazione colonne
                new_value = _build_b27(header, righe_stipendi)

            if new_value != cell.value:
                cell.value = new_value

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# S3 helpers
# ---------------------------------------------------------------------------

def load_template() -> bytes:
    obj = s3.get_object(Bucket=TEMPLATE_BUCKET, Key=TEMPLATE_KEY)
    return obj["Body"].read()


def save_output(id_pratica: str, excel_bytes: bytes) -> str:
    key = f"{OUTPUT_PREFIX}/{id_pratica}/prospetto_{id_pratica}.xlsx"
    s3.put_object(
        Bucket=OUTPUT_BUCKET,
        Key=key,
        Body=excel_bytes,
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ContentDisposition=f'attachment; filename="prospetto_{id_pratica}.xlsx"',
    )
    url = s3.generate_presigned_url(
        "get_object",
        Params={"Bucket": OUTPUT_BUCKET, "Key": key},
        ExpiresIn=PRESIGNED_URL_EXPIRY,
    )
    return url


# ---------------------------------------------------------------------------
# Response helpers
# ---------------------------------------------------------------------------

def response(status_code: int, body: dict) -> dict:
    return {
        "statusCode": status_code,
        "headers": CORS_HEADERS,
        "body": json.dumps(body, ensure_ascii=False, default=str),
    }


# ---------------------------------------------------------------------------
# Handler
# ---------------------------------------------------------------------------

def lambda_handler(event: dict, context) -> dict:
    logger.info("EVENT: %s", json.dumps(event, default=str))

    # ------------------------------------------------------------------ #
    # Gestione evento S3 (trigger da extraction)
    # ------------------------------------------------------------------ #
    if "Records" in event:
        for record in event.get("Records", []):
            if record.get("eventSource") != "aws:s3":
                continue
            
            key = record["s3"]["object"]["key"]
            # Estrai id_pratica dalla key: output/ready/stipendi/{id_pratica}/_trigger.json
            parts = key.split("/")
            if len(parts) >= 4 and parts[0] == "output" and parts[1] == "ready" and parts[2] == "stipendi":
                id_pratica = parts[3]
                logger.info("S3 trigger received for pratica: %s", id_pratica)
                
                # Genera il prospetto
                try:
                    result = _generate_prospetto(id_pratica)
                    logger.info("Prospetto generated from S3 trigger: %s", result.get("url", "").split("?")[0] if result.get("url") else "N/A")
                except Exception as e:
                    logger.error("Error generating prospetto from S3 trigger: %s", str(e))
                    continue
        
        return {"statusCode": 200, "body": json.dumps({"message": "Processing complete"})}
    
    # ------------------------------------------------------------------ #
    # Gestione chiamata HTTP (API Gateway)
    # ------------------------------------------------------------------ #
    http_method = event.get("httpMethod") or (
        event.get("requestContext", {}).get("http", {}).get("method", "")
    )
    if http_method == "OPTIONS":
        return response(200, {})

    path_params = event.get("pathParameters") or {}
    id_pratica = path_params.get("id_pratica", "").strip()

    if not id_pratica:
        return response(400, {"error": "id_pratica mancante nel path"})

    logger.info("Generating prospetto for pratica: %s", id_pratica)
    
    try:
        result = _generate_prospetto(id_pratica)
        return response(200, result)
    except Exception as e:
        logger.exception("Error generating prospetto")
        return response(500, {"error": str(e)})


def _generate_prospetto(id_pratica: str) -> dict:
    """
    Genera il prospetto per una pratica.
    
    Args:
        id_pratica: ID della pratica
    
    Returns:
        Dict con url e id_pratica
    
    Raises:
        Exception: In caso di errore
    """
    # 1. Leggi METADATA
    metadata = query_metadata(id_pratica)
    if not metadata:
        raise ValueError(f"Pratica {id_pratica} non trovata")

    # 2. Leggi decreto di ricostruzione
    decreto = query_documento_decreto(id_pratica)
    if not decreto:
        raise ValueError("Nessun documento decreto_ricostruzione trovato per questa pratica")

    # 3. Leggi visto (opzionale)
    try:
        visto = query_documento_visto(id_pratica)
    except ClientError as e:
        logger.warning("Impossibile recuperare visto: %s", e.response['Error']['Message'])
        visto = None

    # 4. Costruisci mappa placeholder
    mapping = _build_placeholder_map(decreto, metadata, visto)

    logger.info(
        "Placeholder map keys: %s",
        [k for k in mapping if not k.startswith("_")]
    )

    # 5. Carica template
    template_bytes = load_template()

    # 6. Compila Excel
    excel_bytes = fill_excel(template_bytes, mapping)

    # 7. Salva su S3 e restituisci URL
    url = save_output(id_pratica, excel_bytes)

    logger.info("Prospetto generato: %s", url.split("?")[0])
    return {"url": url, "id_pratica": id_pratica}

