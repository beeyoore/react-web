"""
Verifica il contenuto del file Excel generato, in particolare le righe di variazione assegni in B34.
"""
import sys
import pathlib

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl non installato. Installalo con: pip install openpyxl")
    sys.exit(1)

OUTPUT_PATH = pathlib.Path(__file__).parent / "prospetto_test.xlsx"

if not OUTPUT_PATH.exists():
    print(f"File non trovato: {OUTPUT_PATH}")
    sys.exit(1)

wb = load_workbook(OUTPUT_PATH)
ws = wb["Prospetto applicazione"]

print("=" * 80)
print("CONTENUTO CELLA B27 (Variazione Stipendi):")
print("=" * 80)
print(ws["B27"].value)
print()

print("=" * 80)
print("CONTENUTO CELLA B34 (Variazione Assegni):")
print("=" * 80)
print(ws["B34"].value)
print()
